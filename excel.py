import random
import string
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side

import random

import random

def generate_random_data():
    wallet = "pulsetrackersol.com"
    
    # Win rate between 52% and 80% for more realism
    win_rate = random.uniform(42, 80)
    
    # ROI correlated with win rate
    roi = random.uniform(10, 30) + (win_rate - 42) * 5
    
    # Buy amount between $400 and $30000
    buy_amount = round(random.uniform(400, 30000), 2)
    
    # PNL and sell amount based on ROI
    pnl = round(buy_amount * (roi / 100), 2)
    sell_amount = round(buy_amount + pnl, 2)
    
    # Transactions and unique addresses
    buy_transactions = random.randint(16, 70)
    sell_transactions = random.randint(int(buy_transactions * 0.8), buy_transactions)
    unique_addresses = random.randint(max(10, int(buy_transactions * 0.3)), min(56, int(buy_transactions * 0.8)))
    
    # Tokens held inversely related to transactions
    tokens_held = random.randint(0, max(1, int(10 - buy_transactions / 10)))
    
    # Average buy based on buy amount and transactions
    avg_buy = round(buy_amount / buy_transactions, 2)
    
    # Average win and loss related to ROI and win rate
    avg_win = round(random.uniform(10, 130) + (roi / 10), 2)
    avg_loss = round(random.uniform(10, 45) + ((100 - win_rate) / 20), 2)
    
    return [
        wallet, round(win_rate, 2), round(roi, 2), pnl, buy_amount, sell_amount,
        buy_transactions, sell_transactions, unique_addresses, tokens_held,
        avg_buy, avg_win, avg_loss
    ]

def create_excel_with_random_data(num_wallets):
    wb = Workbook()
    ws = wb.active
    ws.title = "pulsetrackersol.com"

    headers = ["Wallet", "Win Rate %", "ROI %", "PNL $", "Buy Amount $", "Sell Amount $",
               "Buy Transactions", "Sell Transactions", "Unique Addresses", "Tokens Held",
               "Average Buy $", "Average Win %", "Average Loss %"]

    ws.append(headers)

    # Define fill colors
    light_blue_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Center, bold, and add border to headers
    for cell in ws[1]:
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.font = Font(bold=True)
        cell.border = thin_border

    for row in range(2, num_wallets + 2):  # Start from row 2 to skip header
        ws.append(generate_random_data())
        
        # Apply alternating background colors
        fill = light_blue_fill if row % 2 == 0 else white_fill
        for cell in ws[row]:
            cell.fill = fill

    # Set the width of each column to 20
    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 20

    wb.save("PulseTrackerSol_daily_wallets.xlsx")
    print(f"Excel file 'high_performance_wallet_data.xlsx' created with {num_wallets} high-performance wallets and alternating row colors.")

# Ask user for the number of wallets to generate
num_wallets = 30
create_excel_with_random_data(num_wallets)
